import pandas as pd
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

def getpathsectionbyrow (excelfile, sheetname):
    df = pd.read_excel(excelfile, sheet_name=sheetname)
    df.drop(columns=['Remark'], inplace=True)
    sections = df['section'].unique()
    result = {}
    for section in sections:
        section_data = df[df['section'] == section]
        section_dict = {}
        for _, row_data in section_data.iterrows():
            section_dict[row_data['Field (Name in VSCode)']] = row_data['Path (or XPath)']
        result[section] = section_dict
    return result

def getdatabycol(excelfile, sheetname):
    df = pd.read_excel(excelfile, sheet_name=sheetname)
    df = df.astype(str)
    colnames = df.columns
    firstcol_datas = df[colnames[0]]
    result = {}
    for firstcol_data in firstcol_datas:
        firstcol = df[df[colnames[0]] == firstcol_data]
        firstcol_dict = {}
        for colname in colnames:
            firstcol_dict[colname] = firstcol[colname].iloc[0]
        result[firstcol_data] = firstcol_dict
    return result 

def getkeysbyorder (dict):
    keys = list(dict.keys())
    return keys

def createxcel(outputname, sheetname, header):

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheetname

    for col_num, col_name in enumerate (header, start=1):
        ws.cell(row=1, column=col_num, value=col_name)

    wb.save(outputname)

def createsheet(filename, sheetname, header):
    wb = load_workbook(filename)
    ws = wb.create_sheet(sheetname)
    for col_num, col_name in enumerate (header, start=1):
        ws.cell(row=1, column=col_num, value=col_name)

    wb.save(filename)


def getcellPosfromHeader(excel_file, sheetName, header_name, row_index):
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook[str(sheetName)]

    header_row = sheet[1]
    col_index = None
    for cell in header_row:
        if cell.value == header_name:
            col_index = cell.column
            break

    workbook.close()

    if col_index is None:
        raise ValueError(f"Header '{header_name}' not found.")

    col_letter = get_column_letter(col_index)
    cell_address = f"{col_letter}{row_index}"
    return cell_address

def getcellPosfromHeader(excel_file, sheetName, header_name):
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook[str(sheetName)]

    header_row = sheet[1]
    col_index = None
    for cell in header_row:
        if cell.value == header_name:
            col_index = cell.column
            break

    workbook.close()

    if col_index is None:
        raise ValueError(f"Header '{header_name}' not found.")

    col_letter = get_column_letter(col_index)
    cell_address = f"{col_letter}"
    return cell_address

def datatoexcel(data, outputname):
    df = pd.DataFrame(data)
    with pd.ExcelWriter(outputname, engine="openpyxl") as writer:
        for sheetname in df["section"].unique():
            data = df[df["section"] == sheetname]
            data.to_excel(writer, sheet_name=str(sheetname), index=False)

def writeToExcel(excel_file, sheetName, pos, val):
    workbook = openpyxl.load_workbook(excel_file)
    sheet_name = str(sheetName)
    selected_sheet = workbook[sheet_name]   
    selected_sheet[pos] = val
    workbook.save(excel_file)
    workbook.close()

def splittext(path):
    splittext = f"[$"
    return  path.split(splittext)

# def main():
#     datatoexcel(data, excelname)

# if __name__ == "__main__":
#     main()